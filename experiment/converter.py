# -*-coding:utf-8-*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Style, Color, PatternFill
from os.path import join, isfile
from config import OUTPUT_PATH
import csv
import traceback
import logging
logger = logging.getLogger(__name__)


def to_csv(rows, filename):
    '''轉換成csv'''
    try:
        rows = [[i.encode('utf-8') for i in j] for j in rows]
    except BaseException:
        logger.warning(traceback.format_exc())

    filepath = join(OUTPUT_PATH, filename + '.csv')
    logger.debug('generating %s' % filepath)
    f = open(filepath, 'w')
    writer = csv.writer(f)
    writer.writerows(rows)
    f.close()


def to_excel(rows, sheetname, filename):
    '''轉換成excel'''
    logger.debug('creating sheet: %s' % sheetname)
    try:
        rows = [[i.encode('utf-8') for i in j] for j in rows]
    except BaseException:
        logger.warning(traceback.format_exc())

    filepath = join(OUTPUT_PATH, filename + '.xlsx')
    if isfile(filepath):
        wb = load_workbook(filepath)
        ws = wb.create_sheet(title=sheetname)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname
    logger.debug('generating %s' % filepath)

    for col_num, row in enumerate(rows, 1):
        for row_num, col in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = col
            # if col == 'y':
            #     cell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FFFF0000')))
    wb.save(filepath)


def _make_list_even_sized(lists):
    max_len = max(len(lst) for lst in lists)
    for lst in lists:
        comp = max_len - len(lst)
        lst += [None] * comp
    return lists
