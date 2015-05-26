# -*-coding:utf-8-*-
from openpyxl import load_workbook
from config import SAMP_PATH, SAMPLE_CHUNK_POLARITY
from jseg import Jieba
from comparison import build_emolex_counter
from collections import OrderedDict
from converter import to_excel
import csv

# '[ㄅ-ㄩ一-龥]+'


def build_emolex_counter_extra():
    pos, neg = build_emolex_counter()
    with open(SAMPLE_CHUNK_POLARITY, 'rU') as f:
        reader = csv.reader(f, dialect=csv.excel_tab)
        rows = [r[0].decode('big5').split(',') for r in reader]
        rows.pop(0)
        for w, pos_ in rows:
            if pos_ == 'N':
                neg[w] = 1
            elif pos_ == 'P':
                pos[w] = 1
            else:
                raise ValueError('unknown pos: %s' % pos)
    return pos, neg





class SampleExp:

    def __init__(self):
        self.wb = load_workbook(SAMP_PATH)
        sheetnames = self.wb.sheetnames
        sheetnames.remove(u'工作表1')
        self.tar_words = sheetnames

    def read_sheet(self, sheetname):
        sheet = self.wb.get_sheet_by_name(sheetname)
        cols, rows = sheet.columns
        cols = [i.value for i in cols]
        rows = [i.value for i in rows]
        map(lambda x: x.pop(0), [cols, rows])
        return zip(cols, rows)

    def to_dic(self):
        dic = OrderedDict()
        for tw in self.tar_words:
            dic[tw] = self.read_sheet(tw)
        return dic

    def to_dic_with_seg(self):
        jieba = Jieba()
        jieba.add_guaranteed_wordlist(self.tar_words)
        dic = self.to_dic()
        for key, row in dic.iteritems():
            con = []
            for pos, sent in row:
                sent_seg = jieba.seg(sent, POS=False)
                output = (pos, sent, sent_seg)
                con.append(output)
            dic[key] = con
        return dic

    def exp(self):
        pos, neg = build_emolex_counter_extra()
        dic = self.to_dic_with_seg()
        for key, row in dic.iteritems():
            poses, sents, segs = zip(*row)
            vals = [sum(((pos.get(w, 0) * 1.0) + (neg.get(w, 0) * -1.0)) for w in sent) for sent in segs]
            output = zip(poses, sents, segs, vals)
            dic[key] = output
        return dic


def exp2():
    pos, neg = build_emolex_counter_extra()

    puncs = [',', '.', '?', '!', '…', '~', '，', '。', '？', '！', '～', '…', ' ']
    puncs = [i.decode('utf-8') for i in puncs]
    se = SampleExp()
    tar_words = se.tar_words
    dic = se.to_dic_with_seg()
    con = []
    for k, group in dic.iteritems():
        for g in group:
            sent = g[2]
            for idx, word in enumerate(sent):
                if word in tar_words:
                    idx = sent.index(word)
                    truncated_sent = sent[idx:]
                    for idx_, w in enumerate(truncated_sent):
                        if w in puncs:
                            rlimit = idx+idx_
                            truncated_sent = sent[idx:rlimit]
                            break
                    pos_val = sum(pos.get(ww, 0) for ww in truncated_sent)
                    neg_val = sum(neg.get(ww, 0) for ww in truncated_sent)
                    fin_val = pos_val * 1.0 + neg_val * -1.0
                    con.append(fin_val)
                    break # !!!!!!!!!!!!
    print len(con)
    vals_con = zip(*[iter(con)]*10)
    return vals_con


def merge():
    se = SampleExp()
    res1 = se.exp()
    keys = res1.keys()
    vals = res1.values()
    res2 = exp2()
    con_outer = []
    for r1, r2 in zip(vals, res2):
        con = []
        for x1, x2 in zip(r1, r2):
            nlst = list(x1)
            nlst.append(x2)
            nlst = [str(i) if type(i) == float else ' '.join(i) if type(i) == list else i for i in nlst]
            con.append(nlst)
        con_outer.append(con)
    # return OrderedDict(zip(keys, con_outer))
    con_outer = det_acc(con_outer)
    for key, rows in zip(keys, con_outer):
        # rows.insert(0, [u'POS', u'原始句', u'斷詞句', u'A_Bag_of_Word', u'Chunk_Based_way'])
        rows.insert(0, [u'Polarity', u'原始句', u'斷詞句', u'A_Bag_of_Word', u'符合', u'Chunk_Based_way', u'符合'])
        rows = zip(*rows)
        to_excel(rows, key, 'sample_experiment')


def convert_val_to_pos(val):
    val = float(val)
    if val > 0:
        pos = 'P'
    elif val < 0:
        pos = 'N'
    elif val == 0:
        pos = 'NEU'
    return pos


def det_acc(con_outer):
    '''determine accuracy'''
    new_con_outer = []
    for group in con_outer:
        con = []
        calc_abow = []
        calc_cba = []
        for pos, ori_sen, seg_sen, abow, cba in group:
            abow_pos = convert_val_to_pos(abow)
            cba_pos = convert_val_to_pos(cba)
            if abow_pos == pos:
                abow_res = 'y'
            else:
                abow_res = 'n'

            if cba_pos == pos:
                cba_res = 'y'
            else:
                cba_res = 'n'

            calc_abow.append(abow_res)
            calc_cba.append(cba_res)

            output = (pos, ori_sen, seg_sen, abow, abow_res, cba, cba_res)
            con.append(output)

        abow_y_num = calc_abow.count('y')
        abow_n_num = calc_abow.count('n')
        cba_y_num = calc_cba.count('y')
        cba_n_num = calc_cba.count('n')

        abow_format = '%dy%dn' % (abow_y_num, abow_n_num)
        cba_format = '%dy%dn' % (cba_y_num, cba_n_num)
        con.append(('', '', '', '', abow_format, '', cba_format))

        abow_acc = abow_y_num / 10.0 * 100
        cba_acc = cba_y_num / 10.0 * 100
        con.append(('', '', '', u'準確率', '%.f%%' % abow_acc, '', '%.f%%' % cba_acc))

        improve = ((cba_acc - abow_acc) / abow_acc) * 100
        con.append(('', '', '', '', '', '', ''))
        con.append(('', '', '', u'提昇率', '%.2f%%' % improve, '', ''))

        new_con_outer.append(con)
    return new_con_outer



if __name__ == '__main__':
    # se = SampleExp()
    # res = se.exp()
    # vals_con = exp2()
    res = merge()
